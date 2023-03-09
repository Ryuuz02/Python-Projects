import requests
import openpyxl

api_key = "RGAPI-d3c95b61-eff0-4127-8143-dffc4210178a"
parameters = {"api_key": api_key}


def get_puuid(summ_name):
    return requests.get("https://na1.api.riotgames.com/lol/summoner/v4/summoners/by-name/" + summ_name,
                        params=parameters).json()["puuid"]


def get_matches(summ_name="Ryuuz"):
    return requests.get("https://americas.api.riotgames.com/lol/match/v5/matches/by-puuid/" + get_puuid(summ_name) +
                        "/ids", params=parameters)


def request_values(match_id):
    return requests.get("https://americas.api.riotgames.com/lol/match/v5/matches/" + match_id, params=parameters)


def find_me(json_dict, summoner_name="Ryuuz"):
    for player in range(10):
        if json_dict["info"]["participants"][player]["summonerName"] == summoner_name:
            return json_dict["info"]["participants"][player]


def change_cell_value(x, y, input_sheet, new_value):
    input_sheet.cell(row=x, column=y).value = new_value


path = "test workbook.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
count = 2
response = get_matches()
for match in response.json():
    running = True
    while running:
        iteration_value = sheet.cell(row=count, column=1).value
        if iteration_value is not None:
            if iteration_value == match:
                running = False
        else:
            summ_dict = find_me(request_values(match).json())
            change_cell_value(count, 1, sheet, match)
            change_cell_value(count, 2, sheet, summ_dict["championName"])
            change_cell_value(count, 3, sheet, (summ_dict["totalMinionsKilled"] + summ_dict["neutralMinionsKilled"]))
            change_cell_value(count, 4, sheet, summ_dict["kills"])
            change_cell_value(count, 5, sheet, summ_dict["deaths"])
            change_cell_value(count, 6, sheet, summ_dict["assists"])
            change_cell_value(count, 7, sheet, summ_dict["goldEarned"])
            change_cell_value(count, 8, sheet, request_values(match).json()["info"]["gameDuration"] / 86400)
            change_cell_value(count, 9, sheet, summ_dict["visionWardsBoughtInGame"])
            change_cell_value(count, 10, sheet, summ_dict["win"])
            with open("GameLog.txt", "a") as f:
                f.write(str(summ_dict))
            running = False
        count += 1
wb.save(path)
